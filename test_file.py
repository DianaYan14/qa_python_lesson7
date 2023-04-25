import csv
import os
import requests
import xlrd
import time
from openpyxl import load_workbook
from pypdf import PdfReader
from selene import browser
from selenium import webdriver
import zipfile


CURRENT_FILE_PATH = os.path.abspath(__file__)
PROJECT_ROOT_PATH = os.path.dirname(CURRENT_FILE_PATH)
resources = os.path.join(PROJECT_ROOT_PATH, 'resources')
download = os.path.join(PROJECT_ROOT_PATH, 'download')


def test_csv():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    csv_path = os.path.join(resources, 'eggs.csv')
    with open(csv_path, 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(csv_path) as csvfile:
        csvreader = csv.reader(csvfile)
        row_count = sum(1 for row in csvreader)
    assert row_count == 2


def test_download_file():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp
    tmp = os.path.join(download, 'pytest-main.zip')
    if not os.path.exists(download):
        os.mkdir(download)
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": download,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)
    browser.config.driver_options = options
    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(5)
    file_size = os.path.getsize(tmp)
    assert file_size == 1564388
    os.remove(tmp)


def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    tmp = os.path.join(resources, 'selenium_logo.png',)
    r = requests.get(url)
    with open(tmp, 'wb') as file:
        file.write(r.content)
        size = os.path.getsize(tmp)
    assert size == 30803


def test_pdf():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    pdf_path = os.path.join(resources, 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(pdf_path)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print(page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412
    assert text == 'pytest Documentation\nRelease 0.1\nholger krekel, trainer and consultant, https://merlinux.eu/\nJul 14, 2022'


def test_xls():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    xls_path = os.path.join(resources, 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(xls_path)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))
    assert book.nsheets == 1
    assert book.sheet_names() == ['Sheet1']
    assert sheet.ncols == 8
    assert sheet.nrows == 10
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'


def test_xlsx():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    xlsx_path = os.path.join(resources, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'


def test_zip_files():
    # TODO создание теста, который заархивирует файлы в resources
    file_dir = os.listdir(resources)
    with zipfile.ZipFile('test.zip', mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in file_dir:
            add_file = os.path.join(resources, file)
            zf.write(add_file, file)

        dock = []
    with zipfile.ZipFile('test.zip', mode='a') as zf:
        for file in file_dir:
            name = os.path.basename(file)
            dock.append(name)
        assert dock == ['docs-pytest-org-en-latest.pdf', 'eggs.csv', 'file_example_XLSX_50.xlsx', 'file_example_XLS_10.xls', 'hello.zip', 'selenium_logo.png']

